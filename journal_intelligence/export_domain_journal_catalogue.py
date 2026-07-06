import json
from pathlib import Path
from collections import Counter

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import A3, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, PageBreak
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet


INPUT = Path("journal_intelligence/domain_filtered/aros_domain_journals.json")

OUTPUT = Path(
    "Research_Output/Research_Infrastructure/Journal_Intelligence"
)

OUTPUT.mkdir(parents=True, exist_ok=True)


def access_class(journal):
    sources = journal.get("source_registry", [])
    access = journal.get("access", {})
    access_type = str(access.get("access_type", "")).lower()

    if "DOAJ" in sources or access.get("doaj_listed") or access_type == "open":
        return "Open Access"

    if access_type == "hybrid":
        return "Hybrid / Limited Access"

    return "Limited Access"


def join_list(value):
    if isinstance(value, list):
        return ", ".join(str(v) for v in value if v)
    return str(value or "")


def ranking_text(journal):
    ranking = journal.get("ranking", {})
    parts = []

    for key in [
        "sjr_quartile",
        "sjr_score",
        "abdc_rating",
        "ugc_care_status"
    ]:
        value = ranking.get(key)
        if value not in ["", None, False]:
            parts.append(f"{key}: {value}")

    return "; ".join(parts)


def main():
    data = json.loads(INPUT.read_text(encoding="utf-8"))
    journals = data["journals"]

    rows = []

    for j in journals:
        rows.append({
            "Journal Name": j.get("journal_name", ""),
            "ISSN": j.get("issn", ""),
            "EISSN": j.get("eissn", ""),
            "Category / AROS Domain": join_list(j.get("aros_domains", [])),
            "Subject Category": join_list(j.get("subject_categories", [])),
            "Publisher": j.get("publisher", ""),
            "Country": j.get("country", ""),
            "Ranking": ranking_text(j),
            "SJR Quartile": j.get("ranking", {}).get("sjr_quartile", ""),
            "ABDC Rating": j.get("ranking", {}).get("abdc_rating", ""),
            "UGC CARE": j.get("ranking", {}).get("ugc_care_status", ""),
            "Source Registry": join_list(j.get("source_registry", [])),
            "Access Classification": access_class(j),
            "URL": j.get("urls", {}).get("journal_url", "")
        })

    headers = list(rows[0].keys())

    xlsx_path = OUTPUT / "AROS_4959_Full_Journal_Catalogue.xlsx"
    pdf_path = OUTPUT / "AROS_4959_Full_Journal_Catalogue.pdf"

    wb = Workbook()

    def write_sheet(name, filtered_rows):
        ws = wb.create_sheet(name)
        ws.append(headers)

        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(wrap_text=True)

        for r in filtered_rows:
            ws.append([r[h] for h in headers])

        ws.freeze_panes = "A2"

        widths = {
            "A": 45, "B": 18, "C": 18, "D": 35, "E": 40,
            "F": 30, "G": 18, "H": 45, "I": 15, "J": 15,
            "K": 15, "L": 25, "M": 24, "N": 45
        }

        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.remove(wb.active)

    write_sheet("Master_All_4959", rows)
    write_sheet(
        "Open_Access",
        [r for r in rows if r["Access Classification"] == "Open Access"]
    )
    write_sheet(
        "Limited_Access",
        [r for r in rows if r["Access Classification"] == "Limited Access"]
    )
    write_sheet(
        "Hybrid_Limited",
        [r for r in rows if r["Access Classification"] == "Hybrid / Limited Access"]
    )

    summary = wb.create_sheet("Summary")
    access_counter = Counter(r["Access Classification"] for r in rows)
    domain_counter = Counter()

    for r in rows:
        for d in r["Category / AROS Domain"].split(","):
            if d.strip():
                domain_counter[d.strip()] += 1

    summary.append(["Metric", "Count"])
    summary.append(["Total Journals", len(rows)])

    for k, v in access_counter.items():
        summary.append([k, v])

    summary.append(["", ""])

    for k, v in domain_counter.most_common():
        summary.append([k, v])

    for cell in summary[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")

    summary.column_dimensions["A"].width = 45
    summary.column_dimensions["B"].width = 15

    wb.save(xlsx_path)

    styles = getSampleStyleSheet()
    doc = SimpleDocTemplate(
        str(pdf_path),
        pagesize=landscape(A3),
        rightMargin=18,
        leftMargin=18,
        topMargin=18,
        bottomMargin=18
    )

    story = []
    story.append(Paragraph("AROS Full Journal Catalogue - 4,959 Journals", styles["Title"]))
    story.append(Paragraph("Includes journal name, ISSN, category, publisher, ranking, source registry, and access classification.", styles["Normal"]))

    story.append(Paragraph("Open Access and Limited Access Classification Summary", styles["Heading2"]))

    summary_table = [["Access Classification", "Count"]]
    for k, v in access_counter.items():
        summary_table.append([k, str(v)])

    table = Table(summary_table, colWidths=[260, 100])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold")
    ]))
    story.append(table)
    story.append(PageBreak())

    pdf_headers = [
        "Journal Name",
        "ISSN",
        "Category / AROS Domain",
        "Publisher",
        "Ranking",
        "Access Classification"
    ]

    for section_name, section_rows in [
        ("OPEN ACCESS JOURNALS", [r for r in rows if r["Access Classification"] == "Open Access"]),
        ("LIMITED ACCESS JOURNALS", [r for r in rows if r["Access Classification"] == "Limited Access"]),
        ("HYBRID / LIMITED ACCESS JOURNALS", [r for r in rows if r["Access Classification"] == "Hybrid / Limited Access"]),
    ]:
        story.append(Paragraph(section_name, styles["Heading1"]))

        table_data = [pdf_headers]

        for r in section_rows:
            table_data.append([
                Paragraph(str(r[h])[:120], styles["Normal"])
                for h in pdf_headers
            ])

        t = Table(
            table_data,
            repeatRows=1,
            colWidths=[230, 80, 180, 150, 180, 110]
        )

        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 6),
            ("VALIGN", (0, 0), (-1, -1), "TOP")
        ]))

        story.append(t)
        story.append(PageBreak())

    doc.build(story)

    print("✓ Full journal catalogue exported")
    print("Excel:", xlsx_path)
    print("PDF:", pdf_path)
    print("Total journals:", len(rows))
    print("Access:", dict(access_counter))


if __name__ == "__main__":
    main()
